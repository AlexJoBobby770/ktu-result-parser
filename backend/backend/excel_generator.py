import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict


def generate_excel_report(students: List[Dict], output_path: str):
    """
    Generate Excel with Name column EMPTY for manual entry
    """
    
    # Convert to flat rows
    flat_data = []
    for student in students:
        regno = student['register_no']
        name = student.get('name', '')  # Empty by default
        dept = student['department']
        
        for subject_code, grade in student['subjects'].items():
            flat_data.append({
                'Register No': regno,
                'Name': name,  # ✅ EMPTY column for manual entry
                'Department': dept,
                'Subject Code': subject_code,
                'Grade': grade,
                'Status': 'Pass' if grade not in ['F', 'FE', 'AB', 'Absent', 'Withheld'] else 'Fail'
            })
    
    df = pd.DataFrame(flat_data)
    
    # Create Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # One sheet per department
        for dept in sorted(df['Department'].unique()):
            dept_df = df[df['Department'] == dept].copy()
            sheet_name = dept.replace('[Full Time]', '').strip()[:31]
            dept_df = dept_df.sort_values(['Register No', 'Subject Code'])
            dept_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Add summary
            worksheet = writer.sheets[sheet_name]
            last_row = len(dept_df) + 3
            
            total_students = dept_df['Register No'].nunique()
            failed_students = dept_df[dept_df['Status'] == 'Fail']['Register No'].nunique()
            total_records = len(dept_df)
            passed_records = len(dept_df[dept_df['Status'] == 'Pass'])
            
            worksheet[f'A{last_row}'] = 'SUMMARY'
            worksheet[f'A{last_row}'].font = Font(bold=True, size=12)
            worksheet[f'A{last_row + 1}'] = 'Total Students:'
            worksheet[f'B{last_row + 1}'] = total_students
            worksheet[f'A{last_row + 2}'] = 'Students with Arrears:'
            worksheet[f'B{last_row + 2}'] = failed_students
            worksheet[f'A{last_row + 3}'] = 'Pass Rate:'
            worksheet[f'B{last_row + 3}'] = f"{round((passed_records / total_records) * 100, 2)}%"
    
    # Styling
    wb = load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Header
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 3, 50)
        
        # Make Name column WIDER for manual entry
        ws.column_dimensions['B'].width = 35  # ✅ Wide for writing names
        
        # Red highlighting for failed grades
        for row in ws.iter_rows(min_row=2, max_col=6):
            if len(row) > 5 and row[5].value == 'Fail':
                row[4].font = Font(color="FF0000", bold=True)
    
    wb.save(output_path)
    print(f"✅ Excel generated: {output_path}")