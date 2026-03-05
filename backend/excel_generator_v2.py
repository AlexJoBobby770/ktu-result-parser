# backend/excel_generator_v2.py (FINAL VERSION)
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict
from collections import defaultdict
from datetime import datetime


# ==================== GRADE & SGPA CALCULATION ====================

GRADE_POINTS = {
    'S': 10, 'A+': 9, 'A': 8.5, 'B+': 8, 'B': 7,
    'C+': 6.5, 'C': 6, 'D': 5.5, 'P': 5,
    'F': 0, 'FE': 0, 'Absent': 0, 'AB': 0, 'Withheld': 0
}

DEFAULT_CREDITS = {
    'MCN401': 0,   # Non-credit
    'MCN202': 0,   # Non-credit
    'CST401': 3,
    'CST423': 3,
    'CST433': 3,
    'ECT435': 3,
    'CSL411': 2,
    'CSQ413': 2,
    'CSD415': 2,
}

FAIL_GRADES = {'F', 'FE', 'Absent', 'AB', 'Withheld'}


def is_skipped_elective(internal_mark: int, grade: str) -> bool:
    """Detect electives that were not chosen (0 internal + Fail grade)"""
    return internal_mark == 0 and grade in FAIL_GRADES


def compute_sgpa(subject_grades: Dict[str, str], internal_marks: Dict[str, int]) -> float:
    """
    Calculate SGPA for a student
    - Skips electives not chosen (internal=0 and grade=F/Absent)
    - Includes failed subjects with 0 grade points
    """
    weighted = 0.0
    total_credits = 0

    for code, grade in subject_grades.items():
        internal = internal_marks.get(code, 0)
        
        # Skip elective not chosen
        if is_skipped_elective(internal, grade):
            continue
        
        # Get grade points (0 for failures)
        gp = GRADE_POINTS.get(grade, 0)
        
        # Get credits (default 3 if not specified)
        credits = DEFAULT_CREDITS.get(code, 3)
        
        weighted += gp * credits
        total_credits += credits
    
    return round(weighted / total_credits, 2) if total_credits > 0 else 0.0


# ==================== MAIN EXCEL GENERATOR ====================

def generate_merged_excel(merged_records: List, output_path: str):
    """
    Generate beautiful Excel with:
    - Department-wise sheets with internal marks + grades + SGPA
    - Subject analysis sheet (FIXED: only counts students who appeared)
    - Status column (removed redundant Arrears column)
    - Proper formatting
    """
    
    COLLEGE_NAME = "ALBERTIAN INSTITUTE OF SCIENCE AND TECHNOLOGY (AISAT)"
    COLLEGE_LOCATION = "Kalamassery, Kerala"
    
    # Convert to DataFrame
    df = pd.DataFrame([r.to_dict() for r in merged_records])
    df = df.sort_values(['Register No', 'Subject Code'])
    
    departments = sorted(df['Department'].unique())
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # ==================== DEPARTMENT SHEETS ====================
        for dept in departments:
            dept_df = df[df['Department'] == dept].copy()
            students = sorted(dept_df['Register No'].unique())
            subjects = sorted(dept_df['Subject Code'].unique())
            
            dept_rows = []
            
            for regno in students:
                student_data = dept_df[dept_df['Register No'] == regno].iloc[0]
                
                row = {
                    'Register No': regno,
                    'Student Name': student_data['Student Name']
                }
                
                grade_dict = {}
                internal_dict = {}
                arrear_list = []
                
                for code in subjects:
                    rec = dept_df[
                        (dept_df['Register No'] == regno) &
                        (dept_df['Subject Code'] == code)
                    ]
                    
                    if not rec.empty:
                        r = rec.iloc[0]
                        internal = r['Internal Mark']
                        grade = r['Grade']
                        subject_name = r['Subject Name']
                        
                        # Check if elective was skipped
                        if is_skipped_elective(internal, grade):
                            row[f"{subject_name} - Internal"] = "—"
                            row[f"{subject_name} - Grade"] = "NOT ELECTED"
                        else:
                            row[f"{subject_name} - Internal"] = internal
                            row[f"{subject_name} - Grade"] = grade
                            
                            grade_dict[code] = grade
                            internal_dict[code] = internal
                            
                            # Track arrears (real failures, not skipped electives)
                            if grade in FAIL_GRADES:
                                arrear_list.append(code)
                    else:
                        row[f"{code} - Internal"] = '—'
                        row[f"{code} - Grade"] = '—'
                
                # Calculate SGPA
                row['SGPA'] = compute_sgpa(grade_dict, internal_dict)
                
                # Status column (removed redundant Arrears column)
                if len(arrear_list) == 0:
                    row['Status'] = '✓ PASS'
                else:
                    row['Status'] = f'✗ {len(arrear_list)} ARREAR(S)'
                
                dept_rows.append(row)
            
            # Write to Excel (start at row 6 to leave space for header)
            pd.DataFrame(dept_rows).to_excel(
                writer, 
                sheet_name=dept, 
                index=False, 
                startrow=5
            )
        
        # ==================== SUBJECT ANALYSIS SHEET (FIXED) ====================
        subject_stats = []
        
        for code in sorted(df['Subject Code'].unique()):
            sd = df[df['Subject Code'] == code]
            
            # FIXED: Only count students who actually appeared for this subject
            # Filter out students who skipped this elective (internal=0 AND grade=F/Absent)
            appeared_students = [
                (row['Internal Mark'], row['Grade'])
                for _, row in sd.iterrows()
                if not is_skipped_elective(row['Internal Mark'], row['Grade'])
            ]
            
            appeared = len(appeared_students)
            
            # Count passed students among those who appeared
            passed = sum(1 for internal, grade in appeared_students if grade not in FAIL_GRADES)
            failed = appeared - passed
            pass_pct = round(passed / appeared * 100, 2) if appeared > 0 else 0
            
            # Calculate average internal (only for students who appeared)
            valid_internals = [internal for internal, grade in appeared_students]
            avg_internal = round(sum(valid_internals) / len(valid_internals), 2) if valid_internals else 0
            
            # Performance rating
            if pass_pct >= 90:
                performance = 'Excellent'
            elif pass_pct >= 75:
                performance = 'Good'
            elif pass_pct >= 60:
                performance = 'Average'
            else:
                performance = 'Needs Improvement'
            
            subject_stats.append({
                'Subject Code': code,
                'Subject Name': sd['Subject Name'].iloc[0],
                'Faculty': sd['Faculty Name'].iloc[0],
                'Appeared': appeared,  # Only students who appeared
                'Passed': passed,
                'Failed': failed,
                'Pass %': pass_pct,
                'Avg Internal': avg_internal,
                'Performance': performance
            })
        
        pd.DataFrame(subject_stats).to_excel(
            writer, 
            sheet_name='Subject Analysis', 
            index=False, 
            startrow=5
        )
    
    # ==================== APPLY FORMATTING ====================
    apply_formatting(output_path, COLLEGE_NAME, COLLEGE_LOCATION, df)
    
    print(f"✅ Excel generated: {output_path}")
    print(f"   📊 {len(merged_records)} records, {df['Register No'].nunique()} students")


# ==================== FORMATTING FUNCTION ====================

def apply_formatting(excel_path: str, college_name: str, location: str, df: pd.DataFrame):
    """Apply professional formatting to all sheets"""
    
    wb = load_workbook(excel_path)
    
    # Color scheme
    colors = {
        'header_blue': '1E3A8A',
        'header_light': '3B82F6',
        'pass_green': 'D1FAE5',
        'fail_red': 'FEE2E2',
        'white': 'FFFFFF',
        'light_gray': 'F3F4F6',
        'dark_text': '1F2937',
    }
    
    # Fonts
    college_font = Font(name='Calibri', size=16, bold=True, color=colors['white'])
    college_fill = PatternFill(start_color=colors['header_blue'], end_color=colors['header_blue'], fill_type='solid')
    
    title_font = Font(name='Calibri', size=12, bold=True, color=colors['dark_text'])
    title_fill = PatternFill(start_color=colors['header_light'], end_color=colors['header_light'], fill_type='solid')
    
    header_font = Font(name='Calibri', size=11, bold=True, color=colors['white'])
    header_fill = PatternFill(start_color=colors['header_blue'], end_color=colors['header_blue'], fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    pass_font = Font(name='Calibri', size=10, color='065F46', bold=True)
    pass_fill = PatternFill(start_color=colors['pass_green'], end_color=colors['pass_green'], fill_type='solid')
    
    fail_font = Font(name='Calibri', size=10, color='991B1B', bold=True)
    fail_fill = PatternFill(start_color=colors['fail_red'], end_color=colors['fail_red'], fill_type='solid')
    
    cell_alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Get the last column with data
        max_col = ws.max_column
        max_col_letter = get_column_letter(max_col)
        
        # ==================== HEADER ROWS ====================
        
        # Row 1: College Name (merge across all columns)
        ws.merge_cells(f'A1:{max_col_letter}1')
        ws['A1'] = college_name
        ws['A1'].font = college_font
        ws['A1'].fill = college_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Row 2: Location (merge across all columns)
        ws.merge_cells(f'A2:{max_col_letter}2')
        ws['A2'] = location
        ws['A2'].font = Font(name='Calibri', size=11, italic=True, color=colors['dark_text'])
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 20
        
        # Row 3: Title (merge across all columns)
        ws.merge_cells(f'A3:{max_col_letter}3')
        if sheet_name == 'Subject Analysis':
            ws['A3'] = f"Comprehensive Subject-wise Performance Analysis"
        else:
            ws['A3'] = f"Internal Assessment & University Results - {sheet_name} Department"
        ws['A3'].font = title_font
        ws['A3'].fill = title_fill
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[3].height = 25
        
        # Row 4: Date (merge across all columns)
        ws.merge_cells(f'A4:{max_col_letter}4')
        ws['A4'] = f"Generated on: {datetime.now().strftime('%d %B %Y at %I:%M %p')}"
        ws['A4'].font = Font(name='Calibri', size=9, italic=True, color='6B7280')
        ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[4].height = 18
        
        # Row 5: Blank
        ws.row_dimensions[5].height = 8
        
        # Row 6: Column Headers
        header_row = 6
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            if cell.value:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
        
        ws.row_dimensions[header_row].height = 35
        
        # ==================== COLUMN WIDTHS ====================
        
        ws.column_dimensions['A'].width = 18  # Register No
        ws.column_dimensions['B'].width = 25  # Student Name
        
        # For remaining columns, set default width
        for col_idx in range(3, max_col + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15
        
        # ==================== DATA ROW FORMATTING ====================
        
        for row_idx in range(7, ws.max_row + 1):
            # Alternating row colors
            if row_idx % 2 == 0:
                row_fill = PatternFill(start_color=colors['light_gray'], end_color=colors['light_gray'], fill_type='solid')
            else:
                row_fill = PatternFill(start_color=colors['white'], end_color=colors['white'], fill_type='solid')
            
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if cell.value is not None:
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                    
                    # Apply background color if not already colored
                    if not cell.fill or cell.fill.start_color.rgb == '00000000':
                        cell.fill = row_fill
                    
                    # Highlight specific values
                    if isinstance(cell.value, str):
                        value_upper = cell.value.upper()
                        
                        # Pass status
                        if 'PASS' in value_upper and '✓' in cell.value:
                            cell.font = pass_font
                            cell.fill = pass_fill
                        
                        # Fail status
                        elif 'ARREAR' in value_upper and '✗' in cell.value:
                            cell.font = fail_font
                            cell.fill = fail_fill
                        
                        # Failed grades
                        elif cell.value in ['F', 'FE', 'Absent', 'AB', 'Withheld']:
                            cell.font = Font(color='991B1B', bold=True)
                            cell.fill = fail_fill
                        
                        # Performance ratings
                        elif 'EXCELLENT' in value_upper:
                            cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
                            cell.font = Font(bold=True, color='065F46')
                        elif 'GOOD' in value_upper:
                            cell.fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
                            cell.font = Font(bold=True, color='1E40AF')
                        elif 'AVERAGE' in value_upper:
                            cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                            cell.font = Font(bold=True, color='92400E')
                        elif 'NEEDS IMPROVEMENT' in value_upper:
                            cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                            cell.font = Font(bold=True, color='991B1B')
        
        # Freeze panes
        ws.freeze_panes = 'A7'
    
    wb.save(excel_path)
    print(f"✅ Formatting applied successfully")


if __name__ == "__main__":
    print("Excel Generator V2 - Final Version")
    print("✅ Fixed subject analysis (only counts students who appeared)")
    print("✅ Removed redundant Arrears column")
    print("✅ Beautiful formatting with full-width headers")