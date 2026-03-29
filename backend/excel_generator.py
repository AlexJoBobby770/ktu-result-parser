# backend/excel_generator.py
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from collections import defaultdict
from datetime import datetime

try:
    from backend.models import PASSING_GRADES, compute_sgpa
except ModuleNotFoundError:
    from models import PASSING_GRADES, compute_sgpa

COLLEGE_NAME     = "ALBERTIAN INSTITUTE OF SCIENCE AND TECHNOLOGY (AISAT)"
COLLEGE_LOCATION = "Kalamassery, Ernakulam, Kerala"

C = {
    "dark_blue":  "1E3A8A", "mid_blue":  "3B82F6",
    "pass_bg":    "D1FAE5", "pass_fg":   "065F46",
    "fail_bg":    "FEE2E2", "fail_fg":   "991B1B",
    "white":      "FFFFFF", "light_gray":"F3F4F6",
    "dark_text":  "1F2937", "skip_bg":   "F3F4F6",
    "skip_fg":    "9CA3AF",
}
THIN = Border(
    left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin",  color="D1D5DB"), bottom=Side(style="thin",color="D1D5DB"),
)
TOP10_BG  = "FEF08A"   # bright yellow
TOP10_FG  = "713F12"   # dark amber
RANK_FILL = "EAB308"   # gold for rank badge cell

PERF_COLORS = {
    "Excellent":         ("D1FAE5", "065F46"),
    "Good":              ("DBEAFE", "1E40AF"),
    "Average":           ("FEF3C7", "92400E"),
    "Needs Improvement": ("FEE2E2", "991B1B"),
}


# ── micro helpers ─────────────────────────────────────────────────────────────

def _fill(h):
    return PatternFill(start_color=h, end_color=h, fill_type="solid")

def _font(color=C["dark_text"], bold=False, size=10):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _batch_year(usn: str) -> str:
    m = re.match(r'[A-Z]+(\d{2})', usn)
    return m.group(1) if m else ""

def _filter_batch(records, batch_year: str, field="usn"):
    if not batch_year:
        return records
    return [r for r in records if _batch_year(getattr(r, field)) == batch_year]


# ── sheet header & formatting helpers ────────────────────────────────────────

def _write_header(ws, title: str, ncols: int):
    cl = get_column_letter(ncols)
    meta = [
        (COLLEGE_NAME,     16, True,  C["white"],    C["dark_blue"]),
        (COLLEGE_LOCATION, 11, False, C["dark_text"],None),
        (title,            12, True,  C["dark_text"],C["mid_blue"]),
        (f"Generated: {datetime.now().strftime('%d %B %Y  %I:%M %p')}",
                            9, False, "6B7280",      None),
    ]
    for i, (txt, sz, bold, fg, bg) in enumerate(meta, 1):
        ws.merge_cells(f"A{i}:{cl}{i}")
        c = ws[f"A{i}"]
        c.value = txt
        c.font  = Font(name="Calibri", size=sz, bold=bold, color=fg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        if bg: c.fill = _fill(bg)
        ws.row_dimensions[i].height = 22 if i < 4 else 16
    ws.row_dimensions[5].height = 6


def _fmt_col_headers(ws, row=6):
    ws.row_dimensions[row].height = 34
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=row, column=col)
        if c.value is not None:
            c.font      = Font(name="Calibri", size=10, bold=True, color=C["white"])
            c.fill      = _fill(C["dark_blue"])
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = THIN


def _fmt_result_data(ws, start=7):
    for r in range(start, ws.max_row + 1):
        alt = _fill(C["light_gray"] if r % 2 == 0 else C["white"])
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=r, column=col)
            if c.value is None: continue
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = THIN
            v = str(c.value).strip()
            if   v == "—":                     c.font, c.fill = _font(C["skip_fg"]),       _fill(C["skip_bg"])
            elif "✓" in v:                     c.font, c.fill = _font(C["pass_fg"],True),  _fill(C["pass_bg"])
            elif "✗" in v:                     c.font, c.fill = _font(C["fail_fg"],True),  _fill(C["fail_bg"])
            elif v in {"F","FE","Absent","Withheld"}: c.font, c.fill = _font(C["fail_fg"],True), _fill(C["fail_bg"])
            elif v in PASSING_GRADES:          c.font, c.fill = _font(C["pass_fg"]),       _fill(C["pass_bg"])
            else:                              c.font, c.fill = _font(),                   alt
    ws.freeze_panes = "A7"


def _fmt_analysis_data(ws, start=7):
    """Colour rows based on Performance column (last column)."""
    perf_col = ws.max_column
    for r in range(start, ws.max_row + 1):
        alt      = _fill(C["light_gray"] if r % 2 == 0 else C["white"])
        perf_val = str(ws.cell(row=r, column=perf_col).value or "").strip()
        bg_hex, fg_hex = PERF_COLORS.get(perf_val, (None, C["dark_text"]))
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=r, column=col)
            if c.value is None: continue
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = THIN
            c.font      = _font(fg_hex if (col == perf_col and bg_hex) else C["dark_text"],
                                bold=(col == perf_col and bool(bg_hex)))
            c.fill      = _fill(bg_hex) if (col == perf_col and bg_hex) else alt
        ws.row_dimensions[r].height = 18
    ws.freeze_panes = "A7"


# ── subject analysis builder ──────────────────────────────────────────────────

def _subject_analysis(ext_records, int_records, batch_year) -> pd.DataFrame:
    """
    Returns one row per subject with: code, name, faculty,
    appeared, passed, failed, pass%, performance.
    Only current-batch students, elective skips excluded.
    """
    # Subject metadata from sessional PDF
    meta = {}   # code → (name, faculty)
    for r in int_records:
        if r.course_code not in meta:
            meta[r.course_code] = (r.subject_name, r.faculty_name)

    # Which (usn, code) pairs were NOT elected
    not_elected = {(r.usn, r.course_code) for r in int_records if not r.elected}

    # Only analyse subjects that appear in the sessional PDF
    # and only for students in the same dept as the sessional PDF
    known_codes   = set(meta.keys())
    sessional_usns = {r.usn for r in int_records}

    # Accumulate grades per subject (current batch, elected only, known codes only)
    code_grades = defaultdict(list)
    for r in _filter_batch(ext_records, batch_year):
        if (r.course_code in known_codes
                and r.usn in sessional_usns
                and (r.usn, r.course_code) not in not_elected):
            code_grades[r.course_code].append(r.grade)

    rows = []
    for code in sorted(code_grades):
        grades_raw = code_grades[code]
        # Exclude FE/Absent/Withheld from appeared count (not considered present)
        grades = [g for g in grades_raw if g not in {"FE", "Absent", "Withheld"}]

        appeared = len(grades)
        if appeared == 0:
            continue

        passed = sum(1 for g in grades if g in PASSING_GRADES)
        # AB (Absent) and any non-passing grade here are considered fail
        failed = appeared - passed
        pct = round(passed / appeared * 100, 1)
        name, faculty = meta.get(code, (code, "N/A"))

        if pct >= 90:   perf = "Excellent"
        elif pct >= 75: perf = "Good"
        elif pct >= 60: perf = "Average"
        else:           perf = "Needs Improvement"

        rows.append({
            "Subject Code": code,
            "Subject Name": name,
            "Faculty":      faculty,
            "Appeared":     appeared,
            "Passed":       passed,
            "Failed":       failed,
            "Pass %":       f"{pct}%",
            "Performance":  perf,
        })
    return pd.DataFrame(rows)


# ── MODE 1 ────────────────────────────────────────────────────────────────────

def generate_external_excel(ext_records: list, output_path: str, batch_year: str = ""):
    dept_map = defaultdict(list)
    for r in _filter_batch(ext_records, batch_year) if batch_year else ext_records:
        dept_map[r.department].append(r)

    # Build stats for summary / arrears / charts
    dept_rows, overall_row, arrears_map = _build_stats(
        ext_records, name_mapping=None, batch_year=batch_year
    )

    titles = {}
    special_sheets = set()
    lbl = f" — Batch 20{batch_year}" if batch_year else ""

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for dept in sorted(dept_map):
            recs = dept_map[dept]
            sg   = defaultdict(dict)
            for r in recs: sg[r.usn][r.course_code] = r.grade
            subs = sorted({r.course_code for r in recs})

            rows = []
            for usn in sorted(sg):
                g   = sg[usn]
                row = {"Register No": usn}
                arr = []
                for s in subs:
                    grade = g.get(s, "")
                    row[s] = grade
                    if grade and grade not in PASSING_GRADES: arr.append(s)
                row["SGPA"]   = compute_sgpa(g)
                row["Status"] = "✓ PASS" if not arr else f"✗ {len(arr)} ARREAR(S)"
                rows.append(row)

            pd.DataFrame(rows).to_excel(writer, sheet_name=dept, index=False, startrow=5)
            titles[dept] = f"{dept} — University Examination Results{lbl}"

        # ── New sheets ──
        _college_summary_sheet(writer, dept_rows, overall_row, batch_year)
        titles["College Summary"] = f"College-Level Pass Percentage{lbl}"
        special_sheets.add("College Summary")

        _arrears_summary_sheet(writer, arrears_map, batch_year)
        titles["Arrears Summary"] = f"Arrears Categorisation{lbl}"
        special_sheets.add("Arrears Summary")

    # Post-format all sheets, then add charts
    _post_format(output_path, titles, analysis_sheets=set(),
                 special_sheets=special_sheets)

    wb = load_workbook(output_path)
    _add_dashboard_charts(wb, dept_rows, arrears_map)
    wb.save(output_path)

    print(f"✅ External-only Excel: {output_path}")

    return dept_rows, overall_row, arrears_map


# ── MODE 2 ────────────────────────────────────────────────────────────────────

def generate_merged_excel(
    ext_records:  list,
    int_records:  list,
    name_mapping: dict,
    output_path:  str,
    batch_year:   str = "",
):
    if batch_year:
        ext_records  = _filter_batch(ext_records, batch_year)
        int_records  = _filter_batch(int_records, batch_year)
        name_mapping = {u: n for u, n in name_mapping.items()
                        if _batch_year(u) == batch_year}

    ext_lu  = defaultdict(dict)
    ext_dep = {}
    for r in ext_records:
        ext_lu[r.usn][r.course_code] = r.grade
        ext_dep[r.usn] = r.department

    int_lu = defaultdict(dict)
    for r in int_records:
        int_lu[r.usn][r.course_code] = r

    # USNs in the sessional Excel - filter ALL output to only these students
    sessional_usns = set(int_lu.keys())

    # Filter external records to only students in sessional Excel
    ext_records = [r for r in ext_records if r.usn in sessional_usns]
    ext_dep = {usn: dept for usn, dept in ext_dep.items() if usn in sessional_usns}

    dept_usns = defaultdict(set)
    for usn, dept in ext_dep.items():
        dept_usns[dept].add(usn)

    # Build stats for summary / arrears / charts
    dept_rows, overall_row, arrears_map = _build_stats(
        ext_records, name_mapping=name_mapping, batch_year=batch_year
    )

    titles = {}
    analysis_sheets = set()
    special_sheets  = set()
    lbl = f" — Batch 20{batch_year}" if batch_year else ""

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        for dept in sorted(dept_usns):
            all_dept_usns = dept_usns[dept]
            # Since we already filtered to sessional_usns, all students here have sessional data
            usns = sorted(all_dept_usns)

            subs     = sorted({c for u in usns for c in ext_lu[u]})

            rows = []
            for usn in usns:
                eg  = ext_lu[usn]
                id_ = int_lu.get(usn, {})  # All students have sessional data now
                row = {"Register No": usn, "Name": name_mapping.get(usn, "")}
                arr, sgpa_g = [], {}

                for code in subs:
                    grade = eg.get(code, "")
                    irec = id_.get(code)
                    if irec and not irec.elected:
                        row[f"{code} Internal"] = "—"
                        row[f"{code} Grade"]    = "—"
                        continue
                    row[f"{code} Internal"] = irec.internal_mark if irec else ""
                    row[f"{code} Grade"]    = grade
                    if grade:
                        sgpa_g[code] = grade
                        if grade not in PASSING_GRADES: arr.append(code)

                row["SGPA"]   = compute_sgpa(sgpa_g)
                row["Status"] = "✓ PASS" if not arr else f"✗ {len(arr)} ARREAR(S)"
                rows.append(row)

            pd.DataFrame(rows).to_excel(writer, sheet_name=dept, index=False, startrow=5)
            titles[dept] = f"{dept} — Internal + University Results{lbl}"

        # Subject Analysis sheet
        if int_records:
            df_analysis = _subject_analysis(ext_records, int_records, batch_year)
            if not df_analysis.empty:
                df_analysis.to_excel(
                    writer, sheet_name="Subject Analysis", index=False, startrow=5
                )
                titles["Subject Analysis"] = f"Subject-wise Pass/Fail Analysis{lbl}"
                analysis_sheets.add("Subject Analysis")

        # ── New sheets ──
        _college_summary_sheet(writer, dept_rows, overall_row, batch_year)
        titles["College Summary"] = f"College-Level Pass Percentage{lbl}"
        special_sheets.add("College Summary")

        _arrears_summary_sheet(writer, arrears_map, batch_year)
        titles["Arrears Summary"] = f"Arrears Categorisation{lbl}"
        special_sheets.add("Arrears Summary")

    # Post-format all sheets, then add charts
    _post_format(output_path, titles, analysis_sheets,
                 special_sheets=special_sheets)

    wb = load_workbook(output_path)
    _add_dashboard_charts(wb, dept_rows, arrears_map)
    wb.save(output_path)

    print(f"✅ Merged Excel: {output_path}")

    return dept_rows, overall_row, arrears_map


# ── top-10 highlighter ────────────────────────────────────────────────────────

def _highlight_top10(ws, start=7):
    """
    Find the SGPA column, pick the top 10 rows by value,
    paint every cell in those rows gold, and write '🏆 #N' in the SGPA cell.
    Handles ties: if two students share rank 10, both are highlighted.
    """
    # Find SGPA column index from header row (row 6)
    sgpa_col = None
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(row=6, column=col).value or "").strip() == "SGPA":
            sgpa_col = col
            break
    if sgpa_col is None:
        return

    # Collect (row_index, sgpa_value) for all data rows
    row_sgpas = []
    for r in range(start, ws.max_row + 1):
        val = ws.cell(row=r, column=sgpa_col).value
        if val is None:
            continue
        try:
            row_sgpas.append((r, float(val)))
        except (ValueError, TypeError):
            pass

    if not row_sgpas:
        return

    # Sort descending, find the cutoff value at rank 10
    row_sgpas.sort(key=lambda x: x[1], reverse=True)
    top10_cutoff = row_sgpas[min(9, len(row_sgpas) - 1)][1]

    # Assign ranks (handle ties: same SGPA = same rank)
    rank = 1
    prev_sgpa = None
    row_ranks = {}
    for i, (r, sgpa) in enumerate(row_sgpas):
        if sgpa != prev_sgpa:
            rank = i + 1
        row_ranks[r] = rank
        prev_sgpa = sgpa

    # Paint top-10 rows
    top_rows = {r for r, sgpa in row_sgpas if sgpa >= top10_cutoff}

    for r in top_rows:
        rank_n = row_ranks[r]
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=r, column=col)
            if c.value is None:
                continue
            if col == sgpa_col:
                c.value = f"🏆 #{rank_n}  {c.value}"
                c.font  = Font(name="Calibri", size=10, bold=True, color=TOP10_FG)
                c.fill  = _fill(RANK_FILL)
            else:
                c.font = Font(name="Calibri", size=10, bold=True, color=TOP10_FG)
                c.fill = _fill(TOP10_BG)


# ── college summary builder ───────────────────────────────────────────────────

def _build_stats(ext_records, name_mapping=None, batch_year=""):
    """
    Build per-department and overall college statistics from external records.
    Returns (dept_rows, overall_row, arrears_map).
      dept_rows : list of dicts with keys: Department, Total, Passed, Failed, Pass %
      overall_row : dict with same keys for the college total
      arrears_map : { usn: { 'name': ..., 'dept': ..., 'count': int, 'subjects': [...] } }
    """
    recs = _filter_batch(ext_records, batch_year) if batch_year else ext_records

    # Per-student: department, arrear subjects, name
    stu_dept    = {}
    stu_arrears = defaultdict(list)   # usn → list of failed course codes
    for r in recs:
        stu_dept[r.usn] = r.department
        if r.grade and r.grade not in PASSING_GRADES:
            stu_arrears[r.usn].append(r.course_code)

    # Per-dept stats
    dept_students = defaultdict(set)
    for usn, dept in stu_dept.items():
        dept_students[dept].add(usn)

    dept_rows = []
    for dept in sorted(dept_students):
        usns    = dept_students[dept]
        total   = len(usns)
        passed  = sum(1 for u in usns if len(stu_arrears.get(u, [])) == 0)
        failed  = total - passed
        pct     = round(passed / total * 100, 1) if total else 0.0
        dept_rows.append({
            "Department": dept, "Total Students": total,
            "Passed": passed, "Failed": failed, "Pass %": f"{pct}%",
        })

    total_all  = len(stu_dept)
    passed_all = sum(1 for u in stu_dept if len(stu_arrears.get(u, [])) == 0)
    failed_all = total_all - passed_all
    pct_all    = round(passed_all / total_all * 100, 1) if total_all else 0.0
    overall_row = {
        "Department": "COLLEGE TOTAL", "Total Students": total_all,
        "Passed": passed_all, "Failed": failed_all, "Pass %": f"{pct_all}%",
    }

    # Arrears map
    arrears_map = {}
    for usn in stu_dept:
        failed_subs = stu_arrears.get(usn, [])
        arrears_map[usn] = {
            "name":     (name_mapping or {}).get(usn, ""),
            "dept":     stu_dept[usn],
            "count":    len(failed_subs),
            "subjects": failed_subs,
        }

    return dept_rows, overall_row, arrears_map


def _college_summary_sheet(writer, dept_rows, overall_row, batch_year=""):
    """Write a 'College Summary' sheet with dept-wise and overall pass %."""
    rows = dept_rows + [overall_row]
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="College Summary", index=False, startrow=5)


def _arrears_summary_sheet(writer, arrears_map, batch_year=""):
    """
    Write an 'Arrears Summary' sheet.
    • Summary table: count per arrear bucket (All Clear, 1, 2, 3, 4, 5+)
    • Detailed list: each student with arrear count, subjects, dept
    """
    # Bucket counts
    buckets = defaultdict(int)
    for info in arrears_map.values():
        c = info["count"]
        if   c == 0: buckets["All Clear"] += 1
        elif c == 1: buckets["1 Arrear"]  += 1
        elif c == 2: buckets["2 Arrears"] += 1
        elif c == 3: buckets["3 Arrears"] += 1
        elif c == 4: buckets["4 Arrears"] += 1
        else:        buckets["5+ Arrears"] += 1

    summary_rows = [{"Category": k, "Count": v}
                    for k, v in [("All Clear", buckets.get("All Clear", 0)),
                                 ("1 Arrear",  buckets.get("1 Arrear", 0)),
                                 ("2 Arrears", buckets.get("2 Arrears", 0)),
                                 ("3 Arrears", buckets.get("3 Arrears", 0)),
                                 ("4 Arrears", buckets.get("4 Arrears", 0)),
                                 ("5+ Arrears", buckets.get("5+ Arrears", 0))]]

    # Detailed list (only students with ≥1 arrear, sorted by count desc)
    detail_rows = []
    for usn, info in sorted(arrears_map.items(), key=lambda x: (-x[1]["count"], x[0])):
        if info["count"] == 0:
            continue
        detail_rows.append({
            "Register No":    usn,
            "Name":           info["name"],
            "Department":     info["dept"],
            "Arrear Count":   info["count"],
            "Failed Subjects": ", ".join(info["subjects"]),
        })

    # Write summary table at top, then detail list below
    df_summary = pd.DataFrame(summary_rows)
    df_detail  = pd.DataFrame(detail_rows)

    df_summary.to_excel(writer, sheet_name="Arrears Summary",
                        index=False, startrow=5)
    gap = 5 + len(df_summary) + 3   # 3-row gap
    df_detail.to_excel(writer, sheet_name="Arrears Summary",
                       index=False, startrow=gap)


def _add_dashboard_charts(wb, dept_rows, arrears_map):
    """
    Add a 'Dashboard' sheet with:
      1. Bar chart  — department-wise pass %
      2. Pie chart  — arrears distribution
    """
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = "3B82F6"

    # ── Write dept data for bar chart ─────────────────────────────
    ws["A1"] = "Department"
    ws["B1"] = "Pass %"
    for i, row in enumerate(dept_rows, 2):
        ws[f"A{i}"] = row["Department"]
        ws[f"B{i}"] = float(row["Pass %"].rstrip("%"))

    end_row = 1 + len(dept_rows)

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.title = "Department-wise Pass Percentage"
    bar.y_axis.title = "Pass %"
    bar.y_axis.scaling.min = 0
    bar.y_axis.scaling.max = 100
    bar.style = 10
    bar.width = 20
    bar.height = 12

    cats = Reference(ws, min_col=1, min_row=2, max_row=end_row)
    data = Reference(ws, min_col=2, min_row=1, max_row=end_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.shape = 4
    ws.add_chart(bar, "D1")

    # ── Write arrears bucket data for pie chart ───────────────────
    pie_start = end_row + 3
    buckets = defaultdict(int)
    for info in arrears_map.values():
        c = info["count"]
        if   c == 0: buckets["All Clear"] += 1
        elif c == 1: buckets["1 Arrear"]  += 1
        elif c == 2: buckets["2 Arrears"] += 1
        elif c == 3: buckets["3 Arrears"] += 1
        elif c == 4: buckets["4 Arrears"] += 1
        else:        buckets["5+ Arrears"] += 1

    ws[f"A{pie_start}"] = "Category"
    ws[f"B{pie_start}"] = "Count"
    bucket_order = ["All Clear", "1 Arrear", "2 Arrears", "3 Arrears",
                    "4 Arrears", "5+ Arrears"]
    r = pie_start + 1
    for cat in bucket_order:
        cnt = buckets.get(cat, 0)
        if cnt > 0:
            ws[f"A{r}"] = cat
            ws[f"B{r}"] = cnt
            r += 1

    pie_end = r - 1

    if pie_end >= pie_start + 1:
        pie = PieChart()
        pie.title = "Arrears Distribution"
        pie.style = 10
        pie.width = 16
        pie.height = 12
        cats = Reference(ws, min_col=1, min_row=pie_start + 1, max_row=pie_end)
        data = Reference(ws, min_col=2, min_row=pie_start, max_row=pie_end)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)

        # Data labels with percentages
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCatName = True
        pie.dataLabels.showVal = False

        ws.add_chart(pie, f"D{end_row + 2}")

    # Style the data cells
    for row in ws.iter_rows(min_row=1, max_row=pie_end, max_col=2):
        for cell in row:
            cell.font = _font(bold=(cell.row == 1 or cell.row == pie_start))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN

    # Column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12


# ── post-format pass ──────────────────────────────────────────────────────────

def _post_format(path: str, titles: dict, analysis_sheets: set,
                 special_sheets: set = None):
    """
    Apply formatting to all sheets.
    special_sheets: set of sheet names that need custom formatting
                    (College Summary, Arrears Summary) rather than result-data styling.
    """
    if special_sheets is None:
        special_sheets = set()
    wb = load_workbook(path)
    for sname in wb.sheetnames:
        if sname == "Dashboard":
            continue   # charts sheet — skip header formatting
        ws    = wb[sname]
        title = titles.get(sname, sname)
        _write_header(ws, title, ws.max_column)
        _fmt_col_headers(ws, row=6)

        if sname in analysis_sheets:
            _fmt_analysis_data(ws, start=7)
            for col, w in zip("ABCDEFGH", [13, 36, 30, 11, 10, 10, 10, 20]):
                ws.column_dimensions[get_column_letter(ord(col)-64)].width = w
        elif sname in special_sheets:
            _fmt_analysis_data(ws, start=7)
            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 16
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 14
            ws.column_dimensions["E"].width = 14
        else:
            _fmt_result_data(ws, start=7)
            _highlight_top10(ws, start=7)
            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 26
    wb.save(path)