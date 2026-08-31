# backend/main.py
import os
import uuid
import shutil
import io

from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware

try:
    from backend.pdf_parser import parse_ktu_pdf
    from backend.internal_parser import parse_sessional_excel
    from backend.excel_generator import generate_external_excel, generate_merged_excel
    from backend.models import PASSING_GRADES
except ModuleNotFoundError:
    from pdf_parser import parse_ktu_pdf
    from internal_parser import parse_sessional_excel
    from excel_generator import generate_external_excel, generate_merged_excel
    from models import PASSING_GRADES

app = FastAPI(title="KTU Result Processor API")

ALLOWED_ORIGINS = [o.strip() for o in os.environ.get("ALLOWED_ORIGINS", "*").split(",") if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "data")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)


@app.get("/")
def root():
    return {"status": "ok", "message": "KTU Result Processor API is running", "docs": "/docs"}


@app.get("/health")
def health():
    return {"status": "ok", "message": "Backend is running"}


@app.post("/upload")
async def upload_result(
    pdf_file:      UploadFile = File(...),
    batch_year:    str        = Form(""),
    internal_file: UploadFile = File(None),
):
    batch_year = batch_year.strip()
    if len(batch_year) == 4 and batch_year.isdigit():
        batch_year = batch_year[2:]
    elif len(batch_year) == 2 and batch_year.isdigit():
        pass
    else:
        batch_year = ""

    session_id = uuid.uuid4().hex[:8]

    ext_path = os.path.join(UPLOAD_DIR, f"{session_id}_result.pdf")
    with open(ext_path, "wb") as f:
        shutil.copyfileobj(pdf_file.file, f)

    print(f"\n[{session_id}] Parsing KTU result PDF...")
    external_records = parse_ktu_pdf(ext_path)

    # PASSING_GRADES is now imported at the top — not inside this function
    student_dept    = {}
    student_arrears = {}
    for r in external_records:
        student_dept[r.usn] = r.department
        student_arrears.setdefault(r.usn, 0)
        if r.grade not in PASSING_GRADES and r.grade != "":
            student_arrears[r.usn] += 1

    total_students  = len(student_dept)
    passed_students = sum(1 for v in student_arrears.values() if v == 0)

    excel_filename = f"{session_id}_results.xlsx"
    excel_path     = os.path.join(OUTPUT_DIR, excel_filename)

    if internal_file and internal_file.filename:
        int_path = os.path.join(UPLOAD_DIR, f"{session_id}_sessional.xlsx")
        with open(int_path, "wb") as f:
            shutil.copyfileobj(internal_file.file, f)

        print(f"[{session_id}] Parsing sessional Excel...")
        internal_records, name_mapping, detected_year = parse_sessional_excel(int_path)

        effective_year = batch_year if batch_year else detected_year
        print(f"[{session_id}] Batch year: 20{effective_year}")

        dept_rows, overall_row, arrears_map = generate_merged_excel(
            external_records, internal_records, name_mapping,
            excel_path, batch_year=effective_year
        )
        mode = "merged"
    else:
        dept_rows, overall_row, arrears_map = generate_external_excel(
            external_records, excel_path, batch_year=batch_year
        )
        mode = "external_only"

    print(f"[{session_id}] Done. Mode={mode}")

    # Build arrears breakdown for the JSON response
    arrears_breakdown = {}
    for info in arrears_map.values():
        c = info["count"]
        if   c == 0: key = "all_clear"
        elif c == 1: key = "1_arrear"
        elif c == 2: key = "2_arrears"
        elif c == 3: key = "3_arrears"
        elif c == 4: key = "4_arrears"
        else:        key = "5+_arrears"
        arrears_breakdown[key] = arrears_breakdown.get(key, 0) + 1

    return {
        "status":             "success",
        "session_id":         session_id,
        "message":            "Files processed successfully",
        "total_students":     total_students,
        "passed_students":    passed_students,
        "pass_percentage":    overall_row["Pass %"],
        "dept_stats":         dept_rows,
        "arrears_breakdown":  arrears_breakdown,
        "mode":               mode,
    }


@app.get("/download/{session_id}")
def download_excel(session_id: str):
    excel_path = os.path.join(OUTPUT_DIR, f"{session_id}_results.xlsx")
    if not os.path.exists(excel_path):
        raise HTTPException(status_code=404, detail="File not found.")
    return FileResponse(
        excel_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"KTU_Results_{session_id}.xlsx",
    )


@app.get("/template")
def download_template():
    """
    Serve a pre-formatted Excel template that matches the internal parser's
    expected format for sessional mark sheets.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sessional Marks"

    # ── Header info (Row 1) ──
    ws["A1"] = "Batch & Semester :YYYY-YYYY SN  (e.g. 2022-2026 S7)"
    ws["A1"].font = Font(name="Calibri", size=11, bold=True, color="1E3A8A")
    ws.merge_cells("A1:H1")

    # ── Instructions (Row 2) ──
    ws["A2"] = ("Instructions: Fill in student data below. "
                "Replace SUB001, SUB002... with actual subject codes (e.g. CST401). "
                "Use * for subjects the student did NOT elect.")
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="6B7280")
    ws.merge_cells("A2:H2")

    # ── Header row (Row 3) ──
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    headers = ["Roll No", "Reg no", "Name", "SUB001", "SUB002", "SUB003",
               "SUB004", "SUB005", "Total"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Mark label row (Row 4) ──
    mark_labels = ["", "", "", "Total Mark", "Total Mark", "Total Mark",
                   "Total Mark", "Total Mark", ""]
    for col_idx, label in enumerate(mark_labels, 1):
        c = ws.cell(row=4, column=col_idx, value=label)
        c.font = Font(name="Calibri", size=9, color="9CA3AF")
        c.alignment = Alignment(horizontal="center")

    # ── Example data rows (5-7) ──
    examples = [
        [1, "AIK22CS001", "STUDENT NAME ONE",   45, 42, 38, "*", 44, 169],
        [2, "AIK22CS002", "STUDENT NAME TWO",   40, "*", 35, 41, 39, 155],
        [3, "AIK22CS003", "STUDENT NAME THREE", 48, 46, 44, 42, "*", 180],
    ]
    for row_idx, row_data in enumerate(examples, 5):
        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.alignment = Alignment(horizontal="center")

    # ── Column widths ──
    widths = [10, 16, 28, 12, 12, 12, 12, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Save to bytes buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=KTU_Internal_Marks_Template.xlsx"
        },
    )